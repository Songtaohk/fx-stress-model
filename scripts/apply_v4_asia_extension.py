from __future__ import annotations

import csv
import json
import math
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


PROJECT = Path("/Users/songtaozhang/Documents/fx-stress-model")
SOURCE_DIR = Path("/private/tmp/fx_v4_sources")
V3_WORKBOOK = PROJECT / "Fx_Stress_Model_Data_v3.xlsx"
V4_WORKBOOK = PROJECT / "Fx_Stress_Model_Data_v4.xlsx"
STRESS_TS = PROJECT / "src/data/stressModel.ts"
V4_JSON = PROJECT / "data/Fx_Stress_Model_Data_v4.json"


YEARS = [str(year) for year in range(1965, 2026)]
ECON_COLS = {
    "cn": "CN\nChina",
    "us": "US\nUnited States",
    "eu": "DE/EA\nGermany / Euro Area",
    "gb": "GB\nUnited Kingdom",
    "jp": "JP\nJapan",
    "kr": "KR\nSouth Korea",
    "in": "IN\nIndia",
    "tw": "TW\nTaiwan",
}


def require_xlrd() -> None:
    try:
        import xlrd  # noqa: F401
    except Exception as exc:  # pragma: no cover
        raise SystemExit(
            "xlrd is required to read Cabinet Office .xls files. "
            "Install with: python3 -m pip install --target /private/tmp/codex_pydeps xlrd==2.0.1"
        ) from exc


def read_xls_sheet(path: Path, sheet_name: str):
    require_xlrd()
    import pandas as pd

    return pd.read_excel(path, sheet_name=sheet_name, header=None, engine="xlrd")


def row_values_by_year(df, row_index: int, header_row: int = 5) -> dict[int, float]:
    years = []
    for value in df.iloc[header_row, 1:].tolist():
        if value is None or (isinstance(value, float) and math.isnan(value)):
            continue
        years.append(int(value))
    result: dict[int, float] = {}
    for offset, year in enumerate(years, start=1):
        value = df.iloc[row_index, offset]
        if value is None or value == "－" or value == "...." or (isinstance(value, float) and math.isnan(value)):
            continue
        result[year] = float(value)
    return result


def load_existing_ts_data():
    text = STRESS_TS.read_text()

    def extract_const(name: str):
        marker = f"export const {name} = "
        start = text.index(marker) + len(marker)
        end = text.index(" as const", start)
        return json.loads(text[start:end])

    return text, extract_const("stressData"), extract_const("stressPairComparisons"), extract_const("stressPairLabels")


def replace_const(text: str, name: str, value) -> str:
    marker = f"export const {name} = "
    start = text.index(marker) + len(marker)
    end = text.index(" as const", start)
    return text[:start] + json.dumps(value, ensure_ascii=False, separators=(",", ":")) + text[end:]


def first_header_row(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [str(v) for v in row if v is not None]
        if values and values[0] == "Year":
            return row_index
    raise ValueError(f"Could not find Year header in {ws.title}")


def write_metric_cell(ws, year: int, econ: str, value: float):
    header_row = first_header_row(ws)
    headers = [cell.value for cell in ws[header_row]]
    year_col = next(index + 1 for index, value in enumerate(headers) if value == "Year")
    target_col = next(index + 1 for index, value in enumerate(headers) if value == ECON_COLS[econ])
    for row_index in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row_index, column=year_col).value == year:
            ws.cell(row=row_index, column=target_col).value = round(value, 4)
            return
    raise ValueError(f"Year {year} not found in {ws.title}")


def load_workbook_series(path: Path):
    wb = load_workbook(path, data_only=True)
    series: dict[str, dict[str, list[float | None]]] = {}
    sheet_to_metric = {
        "CA_GDP": "caGdp",
        "Savings_GDP": "savingsGdp",
        "FDI_GDP": "fdiGdp",
        "Portfolio_GDP": "portfolioGdp",
    }
    for sheet_name, metric in sheet_to_metric.items():
        ws = wb[sheet_name]
        header_row = first_header_row(ws)
        headers = [cell.value for cell in ws[header_row]]
        year_col = next(index + 1 for index, value in enumerate(headers) if value == "Year")
        econ_cols = {
            econ: next(index + 1 for index, value in enumerate(headers) if value == label)
            for econ, label in ECON_COLS.items()
        }
        series[metric] = {econ: [None for _ in YEARS] for econ in ECON_COLS}
        for row_index in range(header_row + 1, ws.max_row + 1):
            year = ws.cell(row=row_index, column=year_col).value
            if year is None or str(year) not in YEARS:
                continue
            idx = YEARS.index(str(year))
            for econ, col in econ_cols.items():
                value = ws.cell(row=row_index, column=col).value
                series[metric][econ][idx] = None if value is None else round(float(value), 4)
    return series


def workbook_gdp_usd():
    wb = load_workbook(V3_WORKBOOK, data_only=True, read_only=True)
    ws = wb["Raw_Nominal_GDP_USD"]
    header_row = first_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[header_row - 1]
    year_col = next(index for index, value in enumerate(headers) if value == "Year")
    tw_col = next(index for index, value in enumerate(headers) if value == "TW GDP current USD")
    result = {}
    for row in rows[header_row:]:
        if row[year_col] is not None:
            result[int(row[year_col])] = float(row[tw_col]) if row[tw_col] is not None else None
    return result


def build_v4_series():
    jp_ext_current = read_xls_sheet(SOURCE_DIR / "jp_external_transactions.xls", "C.Y(1)Current")
    jp_ext_capital = read_xls_sheet(SOURCE_DIR / "jp_external_transactions.xls", "C.Y(2)Capital")
    jp_gde = read_xls_sheet(SOURCE_DIR / "jp_national_disposable_income.xls", "Number")

    jp_gdp = row_values_by_year(jp_gde, 35)
    jp_private_consumption = row_values_by_year(jp_gde, 7)
    jp_gov_consumption = row_values_by_year(jp_gde, 14)
    jp_ca_yen = row_values_by_year(jp_ext_current, 51)
    jp_fdi_assets = row_values_by_year(jp_ext_capital, 11)
    jp_fdi_liabilities = row_values_by_year(jp_ext_capital, 20)
    jp_port_assets = row_values_by_year(jp_ext_capital, 15)
    jp_port_liabilities = row_values_by_year(jp_ext_capital, 24)

    jp = {"caGdp": {}, "savingsGdp": {}, "fdiGdp": {}, "portfolioGdp": {}}
    for year in range(1970, 1996):
        if year in jp_ca_yen and year in jp_gdp:
            jp["caGdp"][year] = jp_ca_yen[year] / jp_gdp[year] * 100
        if year in jp_gdp and year in jp_private_consumption and year in jp_gov_consumption:
            jp["savingsGdp"][year] = (jp_gdp[year] - jp_private_consumption[year] - jp_gov_consumption[year]) / jp_gdp[year] * 100
        if year in jp_fdi_assets and year in jp_fdi_liabilities and year in jp_gdp:
            jp["fdiGdp"][year] = (jp_fdi_assets[year] - jp_fdi_liabilities[year]) / jp_gdp[year] * 100
        if year in jp_port_assets and year in jp_port_liabilities and year in jp_gdp:
            jp["portfolioGdp"][year] = (jp_port_assets[year] - jp_port_liabilities[year]) / jp_gdp[year] * 100

    tw_wb_gdp = workbook_gdp_usd()
    tw_hist = load_workbook(SOURCE_DIR / "tw_hist_standard_year.xlsx", data_only=True, read_only=True).active
    tw = {"caGdp": {}, "fdiGdp": {}, "portfolioGdp": {}}

    def tw_row(row_index: int) -> dict[int, float]:
        row = list(tw_hist.iter_rows(min_row=row_index, max_row=row_index, values_only=True))[0]
        header = list(tw_hist.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        result = {}
        for col_index, year in enumerate(header):
            if isinstance(year, int):
                value = row[col_index]
                if isinstance(value, (int, float)):
                    result[year] = float(value)
        return result

    tw_ca = tw_row(5)
    tw_fdi = tw_row(132)
    tw_port = tw_row(143)
    for year in range(1984, 2012):
        gdp = tw_wb_gdp.get(year)
        if not gdp:
            continue
        if year in tw_ca:
            tw["caGdp"][year] = tw_ca[year] * 1_000_000 / gdp * 100
        if year in tw_fdi:
            tw["fdiGdp"][year] = tw_fdi[year] * 1_000_000 / gdp * 100
        if year in tw_port:
            tw["portfolioGdp"][year] = tw_port[year] * 1_000_000 / gdp * 100
    return jp, tw


def add_log_sheet(wb, jp, tw):
    for name in ["V4_Asia_Extension_Log", "KR_TW_Coverage_Check_v4"]:
        if name in wb.sheetnames:
            del wb[name]
    ws = wb.create_sheet("V4_Asia_Extension_Log")
    rows = [
        ["Metric", "Economy", "Years filled", "Source", "Formula", "Boundary note"],
        ["CA_GDP", "JP", "1970-1995", "Japan Cabinet Office ESRI, National Accounts for 1998, External Transactions, Calendar Year", "Surplus of the nation on current transactions / Gross domestic expenditure * 100", "1996+ remains WDI/IMF/OECD chain; check discontinuity around 1995/1996."],
        ["Savings_GDP", "JP", "1970-1995", "Japan Cabinet Office ESRI, National Accounts for 1998, Gross Domestic Expenditure, Calendar Year", "(Gross domestic expenditure - private final consumption - government final consumption) / Gross domestic expenditure * 100", "1996+ remains WDI/OECD chain."],
        ["FDI_GDP", "JP", "1970-1995", "Japan Cabinet Office ESRI, External Transactions, Capital Transactions, Calendar Year", "(Direct investment assets - Direct investment liabilities) / Gross domestic expenditure * 100", "1996+ remains WDI/IMF/OECD chain."],
        ["Portfolio_GDP", "JP", "1970-1995", "Japan Cabinet Office ESRI, External Transactions, Capital Transactions, Calendar Year", "(Securities investment assets - Securities investment liabilities) / Gross domestic expenditure * 100", "1996+ remains WDI/IMF/OECD chain."],
        ["CA_GDP", "TW", "1984-2011", "CBC Historical Data, Balance of Payments, Standard Presentation by Year", "Current account USD mn / DGBAS nominal GDP USD * 100", "2012+ remains existing CBC open-data chain."],
        ["FDI_GDP", "TW", "1984-2011", "CBC Historical Data, Balance of Payments, Standard Presentation by Year", "Direct investment net USD mn / DGBAS nominal GDP USD * 100", "2012+ remains existing CBC open-data chain."],
        ["Portfolio_GDP", "TW", "1984-2011", "CBC Historical Data, Balance of Payments, Standard Presentation by Year", "Portfolio investment net USD mn / DGBAS nominal GDP USD * 100", "2012+ remains existing CBC open-data chain."],
    ]
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    check = wb.create_sheet("KR_TW_Coverage_Check_v4")
    check_rows = [
        ["Economy", "Finding", "Decision"],
        ["KR", "CA/FDI/Portfolio/Savings already cover 1976-2024; CPI/FX reserves/demographics are longer.", "No v4 backfill needed for these indicators."],
        ["KR", "Policy-rate series starts 1999 and 10Y nominal yield starts 2000 in current aligned source.", "Do not substitute call rate or 3Y bond yield unless a separate proxy metric is approved."],
        ["KR", "External debt/GDP starts 1998 in current aligned source.", "Do not backcast without official same-definition external debt stock."],
        ["TW", "CBC BOP historical XLSX provides annual 1984-2023 data and notes BPM6 conversion.", "Backfill BOP-derived CA/FDI/Portfolio from 1984-2011."],
        ["TW", "Credit_Gap remains blank because BIS does not provide Taiwan aligned credit-to-GDP gap.", "Do not derive until quarterly private-sector credit and GDP are approved."],
    ]
    for row in check_rows:
        check.append(row)
    check.freeze_panes = "A2"
    for cell in check[1]:
        cell.font = Font(bold=True)


def recompute_coverage_summary(wb):
    if "Coverage_Summary" not in wb.sheetnames:
        return
    ws = wb["Coverage_Summary"]
    ws.delete_rows(1, ws.max_row)
    ws.append(["Metric", "Economy", "First Year", "Last Year", "Non-empty Years"])
    for sheet_name in [
        "CA_GDP", "NIIP_GDP", "FX_Reserves_GDP", "CPI_Inflation", "Policy_Rate_Nominal", "Real_Policy_Rate",
        "10Y_Nominal_Yield", "TFP", "ULC_proxy", "FDI_GDP", "Portfolio_GDP", "Savings_GDP", "Old_Age_Dependency",
        "Govt_Debt_GDP", "External_Debt_GDP", "Credit_Gap", "Real_Policy_Rate_Z"
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        source = wb[sheet_name]
        try:
            header_row = first_header_row(source)
        except ValueError:
            continue
        headers = [cell.value for cell in source[header_row]]
        year_col = next((index + 1 for index, value in enumerate(headers) if value == "Year"), None)
        if not year_col:
            continue
        for col_index, header in enumerate(headers, start=1):
            if col_index == year_col or header is None:
                continue
            vals = []
            for row_index in range(header_row + 1, source.max_row + 1):
                year = source.cell(row=row_index, column=year_col).value
                value = source.cell(row=row_index, column=col_index).value
                if isinstance(year, int) and value not in (None, ""):
                    vals.append(year)
            if vals:
                ws.append([sheet_name, header, min(vals), max(vals), len(vals)])
    for cell in ws[1]:
        cell.font = Font(bold=True)


def style_white_background(wb):
    white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    black = Font(color="000000")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = white
                if cell.font:
                    cell.font = cell.font.copy(color="000000")
                else:
                    cell.font = black


def update_workbook(jp, tw):
    shutil.copyfile(V3_WORKBOOK, V4_WORKBOOK)
    wb = load_workbook(V4_WORKBOOK)
    for metric, values in jp.items():
        sheet_name = {"caGdp": "CA_GDP", "savingsGdp": "Savings_GDP", "fdiGdp": "FDI_GDP", "portfolioGdp": "Portfolio_GDP"}[metric]
        for year, value in values.items():
            write_metric_cell(wb[sheet_name], year, "jp", value)
    for metric, values in tw.items():
        sheet_name = {"caGdp": "CA_GDP", "fdiGdp": "FDI_GDP", "portfolioGdp": "Portfolio_GDP"}[metric]
        for year, value in values.items():
            write_metric_cell(wb[sheet_name], year, "tw", value)
    add_log_sheet(wb, jp, tw)
    recompute_coverage_summary(wb)
    style_white_background(wb)
    wb.save(V4_WORKBOOK)


def update_ts_and_json(jp, tw):
    text, stress_data, pair_comp, pair_labels = load_existing_ts_data()
    for metric, values in jp.items():
        arr = stress_data[metric]["jp"]
        for year, value in values.items():
            arr[YEARS.index(str(year))] = round(value, 4)
    for metric, values in tw.items():
        arr = stress_data[metric]["tw"]
        for year, value in values.items():
            arr[YEARS.index(str(year))] = round(value, 4)

    for metric, econ_data in stress_data.items():
        if metric not in pair_comp:
            continue
        for pair_key, labels in pair_labels.items():
            base = labels["base"]
            quote = labels["quote"]
            pair_comp[metric][pair_key] = [
                None if econ_data[base][index] is None or econ_data[quote][index] is None
                else round(econ_data[base][index] - econ_data[quote][index], 4)
                for index in range(len(YEARS))
            ]

    text = replace_const(text, "stressData", stress_data)
    text = replace_const(text, "stressPairComparisons", pair_comp)
    STRESS_TS.write_text(text)

    V4_JSON.write_text(json.dumps({
        "version": "Fx_Stress_Model_Data_v4",
        "generatedAt": "2026-05-24",
        "years": YEARS,
        "v4Extension": {
            "jp": {metric: {str(year): round(value, 4) for year, value in values.items()} for metric, values in jp.items()},
            "tw": {metric: {str(year): round(value, 4) for year, value in values.items()} for metric, values in tw.items()},
            "notes": [
                "JP 1970-1995 uses Japan Cabinet Office ESRI historical national accounts / external transactions.",
                "TW 1984-2011 uses CBC historical BOP annual standard presentation and DGBAS GDP denominator.",
                "No interpolation, extrapolation, or proxy substitution was used."
            ]
        }
    }, ensure_ascii=False, indent=2))


def main():
    jp, tw = build_v4_series()
    update_workbook(jp, tw)
    update_ts_and_json(jp, tw)
    print(json.dumps({
        "jp_filled": {metric: [min(values), max(values), len(values)] for metric, values in jp.items()},
        "tw_filled": {metric: [min(values), max(values), len(values)] for metric, values in tw.items()},
        "workbook": str(V4_WORKBOOK),
        "json": str(V4_JSON),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
